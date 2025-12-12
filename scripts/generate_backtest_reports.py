#!/usr/bin/env python3
"""
回测报告生成模块
生成Excel格式的回测分析报告
"""

import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import TZ_UTC9, REPORTS_DIR, SYMBOLS
from scripts.backtest_engine import (
    BacktestEngine, SimpleFollowStrategy, 
    ReversalStrategy, MultiTimeframeStrategy
)


def run_backtest(symbol_name, strategy, initial_capital=10000):
    """
    运行回测
    
    Args:
        symbol_name: 交易对名称
        strategy: 策略对象
        initial_capital: 初始资金
    
    Returns:
        tuple: (engine, metrics)
    """
    engine = BacktestEngine(symbol_name, initial_capital)
    
    # 获取数据
    monthly_patterns = engine.get_monthly_patterns()
    weekly_patterns = engine.get_weekly_patterns()
    
    # 生成信号
    if isinstance(strategy, MultiTimeframeStrategy):
        signals = strategy.generate_signals(monthly_patterns, weekly_patterns)
    else:
        signals = strategy.generate_signals(monthly_patterns)
    
    # 执行交易
    for signal in signals:
        if signal['action'] in ['long', 'short']:
            engine.open_position(
                direction=signal['action'],
                price=signal['price'],
                size=signal['size'],
                date=signal['date'],
                reason=signal['reason']
            )
        
        # 更新权益曲线
        engine.update_equity(signal['price'], signal['date'])
    
    # 平掉最后的持仓
    if engine.position != 0 and signals:
        engine.close_position(signals[-1]['price'], signals[-1]['date'], "回测结束")
    
    # 计算指标
    metrics = engine.calculate_metrics()
    
    return engine, metrics


def generate_excel_report(results, output_path):
    """
    生成Excel回测报告
    
    Args:
        results: 回测结果字典
        output_path: 输出文件路径
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. 策略总览
        overview_data = []
        for symbol, strategies in results.items():
            for strategy_name, data in strategies.items():
                if data['metrics']:
                    overview_data.append({
                        '交易对': symbol,
                        '策略': strategy_name,
                        '初始资金': data['metrics']['initial_capital'],
                        '最终资金': data['metrics']['final_capital'],
                        '总收益': data['metrics']['total_return'],
                        '收益率(%)': data['metrics']['total_return_percent'],
                        '交易次数': data['metrics']['num_trades'],
                        '胜率(%)': data['metrics']['win_rate'],
                        '盈亏比': data['metrics']['profit_factor'],
                        '最大回撤(%)': data['metrics']['max_drawdown_percent'],
                        '夏普比率': data['metrics']['sharpe_ratio']
                    })
        
        df_overview = pd.DataFrame(overview_data)
        df_overview.to_excel(writer, sheet_name='策略总览', index=False)
        
        # 2. 每个交易对和策略的详细数据
        for symbol, strategies in results.items():
            for strategy_name, data in strategies.items():
                sheet_name = f"{symbol}_{strategy_name}"[:31]  # Excel sheet name limit
                
                # 交易明细
                if data['trades']:
                    df_trades = pd.DataFrame(data['trades'])
                    df_trades.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 权益曲线
                if data['equity_curve']:
                    equity_sheet_name = f"{symbol}_{strategy_name}_权益"[:31]
                    df_equity = pd.DataFrame(data['equity_curve'])
                    df_equity.to_excel(writer, sheet_name=equity_sheet_name, index=False)
        
        # 3. 策略对比
        if len(overview_data) > 1:
            comparison_data = []
            for symbol in results.keys():
                row = {'交易对': symbol}
                for strategy_name in results[symbol].keys():
                    metrics = results[symbol][strategy_name]['metrics']
                    if metrics:
                        row[f'{strategy_name}_收益率'] = metrics['total_return_percent']
                        row[f'{strategy_name}_夏普'] = metrics['sharpe_ratio']
                comparison_data.append(row)
            
            df_comparison = pd.DataFrame(comparison_data)
            df_comparison.to_excel(writer, sheet_name='策略对比', index=False)
    
    # 格式化Excel
    format_excel(output_path)
    
    print(f"回测报告已生成: {output_path}")


def format_excel(file_path):
    """格式化Excel文件"""
    wb = load_workbook(file_path)
    
    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 格式化标题行
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)


def main():
    """主函数"""
    print("=" * 60)
    print("回测分析报告生成")
    print("=" * 60)
    
    # 初始化策略
    strategies = {
        '简单跟随': SimpleFollowStrategy(),
        '反转策略': ReversalStrategy(),
        '多周期结合': MultiTimeframeStrategy()
    }
    
    # 运行回测
    results = {}
    
    for symbol_config in SYMBOLS:
        symbol_name = symbol_config['name']
        print(f"\n{'=' * 60}")
        print(f"回测交易对: {symbol_config['display_name']}")
        print('=' * 60)
        
        results[symbol_name] = {}
        
        for strategy_name, strategy in strategies.items():
            print(f"\n运行策略: {strategy_name}")
            
            try:
                engine, metrics = run_backtest(symbol_name, strategy)
                
                if metrics:
                    print(f"  初始资金: ${metrics['initial_capital']}")
                    print(f"  最终资金: ${metrics['final_capital']}")
                    print(f"  总收益: ${metrics['total_return']} ({metrics['total_return_percent']}%)")
                    print(f"  交易次数: {metrics['num_trades']}")
                    print(f"  胜率: {metrics['win_rate']}%")
                    print(f"  最大回撤: {metrics['max_drawdown_percent']}%")
                    print(f"  夏普比率: {metrics['sharpe_ratio']}")
                    
                    results[symbol_name][strategy_name] = {
                        'metrics': metrics,
                        'trades': engine.trades,
                        'equity_curve': engine.equity_curve
                    }
                else:
                    print(f"  无足够数据进行回测")
                    results[symbol_name][strategy_name] = {
                        'metrics': None,
                        'trades': [],
                        'equity_curve': []
                    }
            
            except Exception as e:
                print(f"  回测失败: {e}")
                results[symbol_name][strategy_name] = {
                    'metrics': None,
                    'trades': [],
                    'equity_curve': []
                }
    
    # 生成报告
    excel_dir = os.path.join(REPORTS_DIR, 'excel')
    os.makedirs(excel_dir, exist_ok=True)
    
    timestamp = datetime.now(TZ_UTC9).strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(excel_dir, f'回测分析报告_{timestamp}.xlsx')
    output_path_latest = os.path.join(excel_dir, '回测分析报告_最新.xlsx')
    
    generate_excel_report(results, output_path)
    generate_excel_report(results, output_path_latest)
    
    print("\n" + "=" * 60)
    print("回测报告生成完成")
    print("=" * 60)
    print(f"报告路径: {output_path}")
    
    return True


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)

