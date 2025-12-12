"""
导出所有数据到Excel
包含周数据、月度模式等完整数据
"""

import sqlite3
import os
import sys
from datetime import datetime

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import DATABASE_PATH, REPORTS_DIR, TZ_UTC9

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def style_excel_header(ws, row_num=1):
    """设置Excel表头样式"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def style_data_cells(ws, start_row=2):
    """设置数据单元格样式"""
    data_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    amdx_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    xamd_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border
            
            if cell.value == 'AMDX':
                cell.fill = amdx_fill
            elif cell.value == 'XAMD':
                cell.fill = xamd_fill


def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    for char in str(cell.value):
                        if '\u4e00' <= char <= '\u9fff':
                            cell_length += 1
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def get_xamd_from_monthly_pattern(pattern, week_of_month):
    """
    根据月度模式表的pattern字段和月内第几周
    转换为X/A/M/D走势
    
    pattern是一个4字母字符串（XAMD或AMDX），每个字母对应月内第1-4周
    例如：
    - pattern='XAMD', week_of_month=1 -> 'X'
    - pattern='XAMD', week_of_month=2 -> 'A'
    - pattern='XAMD', week_of_month=3 -> 'M'
    - pattern='XAMD', week_of_month=4 -> 'D'
    - pattern='AMDX', week_of_month=1 -> 'A'
    - pattern='AMDX', week_of_month=2 -> 'M'
    - pattern='AMDX', week_of_month=3 -> 'D'
    - pattern='AMDX', week_of_month=4 -> 'X'
    """
    if not pattern or len(pattern) != 4:
        return ''
    
    # week_of_month是1-4，对应pattern的索引0-3
    if 1 <= week_of_month <= 4:
        return pattern[week_of_month - 1]
    else:
        return ''


def export_all_data(conn):
    """导出所有数据到Excel"""
    print("=" * 60)
    print("导出所有数据到Excel")
    print("=" * 60)
    
    excel_dir = os.path.join(REPORTS_DIR, 'excel')
    os.makedirs(excel_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = os.path.join(excel_dir, f'完整数据导出_{timestamp}.xlsx')
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # ========== 工作表1: 总体汇总 ==========
        print("生成工作表: 总体汇总...")
        query = """
            SELECT 
                s.symbol as '交易对',
                COUNT(*) as '总月数',
                SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
                SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
                ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
                ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)',
                MIN(mp.year || '-' || printf('%02d', mp.month)) as '数据起始月',
                MAX(mp.year || '-' || printf('%02d', mp.month)) as '数据结束月'
            FROM monthly_patterns mp
            JOIN symbols s ON mp.symbol_id = s.id
            GROUP BY s.symbol
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='总体汇总', index=False)
        ws = writer.sheets['总体汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ========== 工作表2: 年度汇总 ==========
        print("生成工作表: 年度汇总...")
        query = """
            SELECT 
                s.symbol as '交易对',
                mp.year as '年份',
                COUNT(*) as '总月数',
                SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
                SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
                ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
                ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)',
                SUM(CASE WHEN mp.is_breakout_up = 1 THEN 1 ELSE 0 END) as '向上突破次数',
                SUM(CASE WHEN mp.is_breakout_down = 1 THEN 1 ELSE 0 END) as '向下突破次数',
                ROUND(AVG(CASE WHEN mp.breakout_up_percent IS NOT NULL THEN mp.breakout_up_percent END), 2) as '平均向上突破幅度(%)',
                ROUND(AVG(CASE WHEN mp.breakout_down_percent IS NOT NULL THEN mp.breakout_down_percent END), 2) as '平均向下突破幅度(%)'
            FROM monthly_patterns mp
            JOIN symbols s ON mp.symbol_id = s.id
            GROUP BY s.symbol, mp.year
            ORDER BY s.symbol, mp.year
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='年度汇总', index=False)
        ws = writer.sheets['年度汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ========== 新增工作表: BTC月份分布统计 ==========
        print("生成工作表: BTC月份分布统计...")
        query = """
            SELECT 
                mp.month as '月份',
                SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
                SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
                COUNT(*) as '总次数',
                ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
                ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)'
            FROM monthly_patterns mp
            JOIN symbols s ON mp.symbol_id = s.id
            WHERE s.symbol = 'BTCUSDT'
            GROUP BY mp.month
            ORDER BY mp.month
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='BTC月份分布统计', index=False)
        ws = writer.sheets['BTC月份分布统计']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ========== 新增工作表: ETH月份分布统计 ==========
        print("生成工作表: ETH月份分布统计...")
        query = """
            SELECT 
                mp.month as '月份',
                SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
                SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
                COUNT(*) as '总次数',
                ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
                ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)'
            FROM monthly_patterns mp
            JOIN symbols s ON mp.symbol_id = s.id
            WHERE s.symbol = 'ETHUSDT'
            GROUP BY mp.month
            ORDER BY mp.month
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='ETH月份分布统计', index=False)
        ws = writer.sheets['ETH月份分布统计']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ========== 按交易对分别创建详细工作表 ==========
        symbols_query = "SELECT id, symbol FROM symbols WHERE is_active = 1"
        symbols_df = pd.read_sql_query(symbols_query, conn)
        
        # 先获取所有月度模式数据，用于映射
        monthly_patterns_query = """
            SELECT 
                symbol_id,
                year,
                month,
                pattern
            FROM monthly_patterns
        """
        monthly_patterns_df = pd.read_sql_query(monthly_patterns_query, conn)
        
        for _, row in symbols_df.iterrows():
            symbol_id = row['id']
            symbol = row['symbol']
            
            print(f"生成工作表: {symbol}_周数据...")
            # 周数据（增加X/A/M/D走势列）
            query = """
                SELECT 
                    week_start as '周开始时间',
                    week_end as '周结束时间',
                    year as '年份',
                    month as '月份',
                    week_of_year as '年内第几周',
                    week_of_month as '月内第几周',
                    week_high as '周最高价',
                    week_low as '周最低价',
                    week_open as '周开盘价',
                    week_close as '周收盘价',
                    data_points as '数据点数',
                    data_quality_score as '数据质量分数'
                FROM weekly_data
                WHERE symbol_id = ?
                ORDER BY week_start
            """
            df = pd.read_sql_query(query, conn, params=(symbol_id,))
            
            # 根据月度模式表计算X/A/M/D走势
            xamd_patterns = []
            for idx, row_data in df.iterrows():
                week_of_month = row_data['月内第几周']
                year = row_data['年份']
                month = row_data['月份']
                
                # 只处理月内第1-4周，第5周及以后显示为空
                if week_of_month > 4:
                    xamd_patterns.append('')
                    continue
                
                # 从月度模式表中查找对应的模式
                monthly_pattern = monthly_patterns_df[
                    (monthly_patterns_df['symbol_id'] == symbol_id) &
                    (monthly_patterns_df['year'] == year) &
                    (monthly_patterns_df['month'] == month)
                ]
                
                if len(monthly_pattern) > 0:
                    mp_row = monthly_pattern.iloc[0]
                    pattern = mp_row['pattern']
                    
                    # 根据月内第几周，从pattern字符串中取对应位置的字母
                    xamd = get_xamd_from_monthly_pattern(pattern, week_of_month)
                    xamd_patterns.append(xamd)
                else:
                    xamd_patterns.append('')  # 没有对应的月度模式数据
            
            # 插入X/A/M/D走势列（在"月内第几周"之后）
            df.insert(df.columns.get_loc('月内第几周') + 1, '走势', xamd_patterns)
            
            sheet_name = f"{symbol}_周数据"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            style_excel_header(ws)
            style_data_cells(ws)
            auto_adjust_column_width(ws)
            
            print(f"生成工作表: {symbol}_月度模式...")
            # 月度模式
            query = """
                SELECT 
                    year as '年份',
                    month as '月份',
                    first_week_start as '第一周开始日期',
                    pattern as '模式',
                    first_week_high as '第一周最高价',
                    first_week_low as '第一周最低价',
                    previous_week_high as '前一周最高价',
                    previous_week_low as '前一周最低价',
                    CASE WHEN is_breakout_up = 1 THEN '是' ELSE '否' END as '向上突破',
                    CASE WHEN is_breakout_down = 1 THEN '是' ELSE '否' END as '向下突破',
                    ROUND(breakout_up_percent, 2) as '向上突破幅度(%)',
                    ROUND(breakout_down_percent, 2) as '向下突破幅度(%)',
                    data_quality_score as '数据质量分数'
                FROM monthly_patterns
                WHERE symbol_id = ?
                ORDER BY year, month
            """
            df = pd.read_sql_query(query, conn, params=(symbol_id,))
            sheet_name = f"{symbol}_月度模式"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            style_excel_header(ws)
            style_data_cells(ws)
            auto_adjust_column_width(ws)
    
    print(f"\n完整数据导出完成: {excel_path}")
    
    # 同时保存最新版本
    latest_path = os.path.join(excel_dir, '完整数据导出_最新.xlsx')
    import shutil
    shutil.copy(excel_path, latest_path)
    print(f"最新版本已保存: {latest_path}")
    
    return excel_path


def main():
    """主函数"""
    conn = sqlite3.connect(DATABASE_PATH)
    
    try:
        export_all_data(conn)
        print("\n" + "=" * 60)
        print("所有数据导出完成!")
        print("=" * 60)
    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        conn.close()


if __name__ == '__main__':
    main()

