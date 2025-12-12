"""
报告生成脚本
生成Excel和PDF格式的分析报告
"""

import sqlite3
import os
import sys
from datetime import datetime

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import DATABASE_PATH, REPORTS_DIR, TZ_UTC9, REPORT_CONFIG

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList


def get_monthly_data(conn):
    """获取月度详细数据"""
    query = """
        SELECT 
            s.symbol as '交易对',
            mp.year as '年份',
            mp.month as '月份',
            mp.first_week_start as '第一周开始日期',
            mp.pattern as '模式',
            mp.first_week_high as '第一周最高价',
            mp.first_week_low as '第一周最低价',
            mp.previous_week_high as '前一周最高价',
            mp.previous_week_low as '前一周最低价',
            CASE WHEN mp.is_breakout_up = 1 THEN '是' ELSE '否' END as '向上突破',
            CASE WHEN mp.is_breakout_down = 1 THEN '是' ELSE '否' END as '向下突破',
            ROUND(mp.breakout_up_percent, 2) as '向上突破幅度(%)',
            ROUND(mp.breakout_down_percent, 2) as '向下突破幅度(%)',
            mp.data_quality_score as '数据质量分数'
        FROM monthly_patterns mp
        JOIN symbols s ON mp.symbol_id = s.id
        ORDER BY s.symbol, mp.year, mp.month
    """
    return pd.read_sql_query(query, conn)


def get_yearly_summary(conn):
    """获取年度汇总数据"""
    query = """
        SELECT 
            s.symbol as '交易对',
            mp.year as '年份',
            COUNT(*) as '总月数',
            SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
            SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
            ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
            ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)',
            SUM(CASE WHEN mp.is_breakout_up = 1 THEN 1 ELSE 0 END) as '向上突破次数',
            SUM(CASE WHEN mp.is_breakout_down = 1 THEN 1 ELSE 0 END) as '向下突破次数',
            ROUND(AVG(CASE WHEN mp.breakout_up_percent IS NOT NULL THEN mp.breakout_up_percent END), 2) as '平均向上突破幅度(%)',
            ROUND(AVG(CASE WHEN mp.breakout_down_percent IS NOT NULL THEN mp.breakout_down_percent END), 2) as '平均向下突破幅度(%)'
        FROM monthly_patterns mp
        JOIN symbols s ON mp.symbol_id = s.id
        GROUP BY s.symbol, mp.year
        ORDER BY s.symbol, mp.year
    """
    return pd.read_sql_query(query, conn)


def get_overall_summary(conn):
    """获取总体汇总数据"""
    query = """
        SELECT 
            s.symbol as '交易对',
            COUNT(*) as '总月数',
            SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
            SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
            ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
            ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)',
            MIN(mp.year || '-' || printf('%02d', mp.month)) as '数据起始月',
            MAX(mp.year || '-' || printf('%02d', mp.month)) as '数据结束月'
        FROM monthly_patterns mp
        JOIN symbols s ON mp.symbol_id = s.id
        GROUP BY s.symbol
    """
    return pd.read_sql_query(query, conn)


def get_pattern_distribution(conn):
    """获取模式分布数据（按月份统计）"""
    query = """
        SELECT 
            mp.month as '月份',
            SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
            SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
            COUNT(*) as '总次数'
        FROM monthly_patterns mp
        GROUP BY mp.month
        ORDER BY mp.month
    """
    return pd.read_sql_query(query, conn)


def style_excel_header(ws, row_num=1):
    """设置Excel表头样式"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def style_data_cells(ws, start_row=2):
    """设置数据单元格样式"""
    data_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # AMDX用绿色，XAMD用红色
    amdx_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    xamd_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 模式列着色
            if cell.value == 'AMDX':
                cell.fill = amdx_fill
            elif cell.value == 'XAMD':
                cell.fill = xamd_fill


def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    # 中文字符算2个宽度
                    for char in str(cell.value):
                        if '\u4e00' <= char <= '\u9fff':
                            cell_length += 1
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_excel_report(conn):
    """生成Excel报告"""
    print("生成Excel报告...")
    
    # 获取数据
    monthly_df = get_monthly_data(conn)
    yearly_df = get_yearly_summary(conn)
    overall_df = get_overall_summary(conn)
    distribution_df = get_pattern_distribution(conn)
    
    # 创建Excel文件
    excel_dir = os.path.join(REPORTS_DIR, 'excel')
    os.makedirs(excel_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = os.path.join(excel_dir, f'AMDX_XAMD_分析报告_{timestamp}.xlsx')
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # 工作表1: 总体汇总
        overall_df.to_excel(writer, sheet_name='总体汇总', index=False)
        ws = writer.sheets['总体汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表2: 年度汇总
        yearly_df.to_excel(writer, sheet_name='年度汇总', index=False)
        ws = writer.sheets['年度汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表3: 月度详细
        monthly_df.to_excel(writer, sheet_name='月度详细', index=False)
        ws = writer.sheets['月度详细']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表4: 模式分布（按月份）
        distribution_df.to_excel(writer, sheet_name='月份分布', index=False)
        ws = writer.sheets['月份分布']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 按交易对分别创建工作表
        symbols = monthly_df['交易对'].unique()
        for symbol in symbols:
            symbol_df = monthly_df[monthly_df['交易对'] == symbol].copy()
            sheet_name = f"{symbol}_详细"[:31]
            symbol_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            style_excel_header(ws)
            style_data_cells(ws)
            auto_adjust_column_width(ws)
    
    print(f"  Excel报告已保存: {excel_path}")
    
    # 同时保存一个最新版本（不带时间戳）
    latest_path = os.path.join(excel_dir, 'AMDX_XAMD_分析报告_最新.xlsx')
    
    with pd.ExcelWriter(latest_path, engine='openpyxl') as writer:
        overall_df.to_excel(writer, sheet_name='总体汇总', index=False)
        ws = writer.sheets['总体汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        yearly_df.to_excel(writer, sheet_name='年度汇总', index=False)
        ws = writer.sheets['年度汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        monthly_df.to_excel(writer, sheet_name='月度详细', index=False)
        ws = writer.sheets['月度详细']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        distribution_df.to_excel(writer, sheet_name='月份分布', index=False)
        ws = writer.sheets['月份分布']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        for symbol in symbols:
            symbol_df = monthly_df[monthly_df['交易对'] == symbol].copy()
            sheet_name = f"{symbol}_详细"[:31]
            symbol_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            style_excel_header(ws)
            style_data_cells(ws)
            auto_adjust_column_width(ws)
    
    print(f"  最新版本已保存: {latest_path}")
    
    return excel_path, latest_path


def generate_pdf_report(conn):
    """生成PDF报告"""
    print("生成PDF报告...")
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        print("  警告: reportlab未安装，跳过PDF生成")
        print("  请运行: pip install reportlab")
        return None
    
    # 获取数据
    monthly_df = get_monthly_data(conn)
    yearly_df = get_yearly_summary(conn)
    overall_df = get_overall_summary(conn)
    
    # 创建PDF
    pdf_dir = os.path.join(REPORTS_DIR, 'pdf')
    os.makedirs(pdf_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    pdf_path = os.path.join(pdf_dir, f'AMDX_XAMD_分析报告_{timestamp}.pdf')
    
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )
    
    # 样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # 居中
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        spaceBefore=20
    )
    
    elements = []
    
    # 标题
    elements.append(Paragraph("AMDX/XAMD Pattern Analysis Report", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # 总体汇总表
    elements.append(Paragraph("Overall Summary", subtitle_style))
    
    # 转换DataFrame为表格数据
    overall_data = [overall_df.columns.tolist()] + overall_df.values.tolist()
    
    overall_table = Table(overall_data)
    overall_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8E8E8')])
    ]))
    elements.append(overall_table)
    elements.append(PageBreak())
    
    # 年度汇总表
    elements.append(Paragraph("Yearly Summary", subtitle_style))
    
    yearly_data = [yearly_df.columns.tolist()] + yearly_df.values.tolist()
    
    yearly_table = Table(yearly_data)
    yearly_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8E8E8')])
    ]))
    elements.append(yearly_table)
    
    # 构建PDF
    doc.build(elements)
    
    print(f"  PDF报告已保存: {pdf_path}")
    
    # 保存最新版本
    latest_path = os.path.join(pdf_dir, 'AMDX_XAMD_分析报告_最新.pdf')
    import shutil
    shutil.copy(pdf_path, latest_path)
    print(f"  最新版本已保存: {latest_path}")
    
    return pdf_path


def export_data_json(conn):
    """导出JSON格式数据（用于GitHub Pages等）"""
    print("导出JSON数据...")
    
    import json
    
    # 获取数据
    monthly_df = get_monthly_data(conn)
    yearly_df = get_yearly_summary(conn)
    overall_df = get_overall_summary(conn)
    
    data_dir = os.path.join(REPORTS_DIR, 'data')
    os.makedirs(data_dir, exist_ok=True)
    
    # 月度数据
    monthly_path = os.path.join(data_dir, 'monthly_patterns.json')
    monthly_df.to_json(monthly_path, orient='records', force_ascii=False, indent=2)
    
    # 年度数据
    yearly_path = os.path.join(data_dir, 'yearly_summary.json')
    yearly_df.to_json(yearly_path, orient='records', force_ascii=False, indent=2)
    
    # 总体数据
    overall_path = os.path.join(data_dir, 'overall_summary.json')
    overall_df.to_json(overall_path, orient='records', force_ascii=False, indent=2)
    
    # 合并数据
    combined = {
        'generated_at': datetime.now(TZ_UTC9).strftime('%Y-%m-%d %H:%M:%S'),
        'timezone': 'UTC+9',
        'overall': overall_df.to_dict(orient='records'),
        'yearly': yearly_df.to_dict(orient='records'),
        'monthly': monthly_df.to_dict(orient='records')
    }
    
    combined_path = os.path.join(data_dir, 'all_data.json')
    with open(combined_path, 'w', encoding='utf-8') as f:
        json.dump(combined, f, ensure_ascii=False, indent=2)
    
    print(f"  JSON数据已保存: {data_dir}")
    
    return combined_path


def main():
    """主函数"""
    print("=" * 60)
    print("AMDX/XAMD 报告生成程序")
    print("=" * 60)
    print(f"当前时间: {datetime.now(TZ_UTC9).strftime('%Y-%m-%d %H:%M:%S')} (UTC+9)")
    
    # 连接数据库
    conn = sqlite3.connect(DATABASE_PATH)
    
    try:
        # 检查数据是否存在
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM monthly_patterns")
        count = cursor.fetchone()[0]
        
        if count == 0:
            print("\n警告: 数据库中没有模式数据，请先运行:")
            print("  1. python scripts/fetch_data.py")
            print("  2. python scripts/calculate_patterns.py")
            return
        
        print(f"\n数据库中有 {count} 条模式记录")
        
        # 生成报告
        generate_excel_report(conn)
        generate_pdf_report(conn)
        export_data_json(conn)
        
        print("\n" + "=" * 60)
        print("报告生成完成!")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        conn.close()


if __name__ == '__main__':
    main()

