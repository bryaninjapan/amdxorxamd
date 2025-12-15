"""
周度模式与走势明细统计脚本
统计A/M/D/X模式在不同走势明细下的出现次数，按年份分组
"""

import sqlite3
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import DATABASE_PATH, REPORTS_DIR

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def style_excel_header(ws, row_num=1):
    """设置Excel表头样式"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def style_data_cells(ws, start_row=2):
    """设置数据单元格样式"""
    data_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border


def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    for char in str(cell.value):
                        if '\u4e00' <= char <= '\u9fff':
                            cell_length += 1
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def get_statistics_data(conn, symbol_name):
    """获取统计数据"""
    query = """
        SELECT 
            dd.year as '年份',
            CASE 
                WHEN wp.pattern IS NULL THEN 'N/A'
                WHEN dd.day_of_week = 0 THEN SUBSTR(wp.pattern, 1, 1)
                WHEN dd.day_of_week = 1 THEN SUBSTR(wp.pattern, 2, 1)
                WHEN dd.day_of_week = 2 THEN SUBSTR(wp.pattern, 3, 1)
                WHEN dd.day_of_week = 3 THEN SUBSTR(wp.pattern, 4, 1)
                WHEN dd.day_of_week = 4 THEN SUBSTR(wp.pattern, 5, 1)
                WHEN dd.day_of_week = 5 THEN SUBSTR(wp.pattern, 6, 1)
                WHEN dd.day_of_week = 6 THEN SUBSTR(wp.pattern, 7, 1)
                ELSE 'N/A'
            END as '周度模式',
            CASE dd.day_of_week
                WHEN 0 THEN COALESCE(wp.monday_trend_detail, 'N/A')
                WHEN 1 THEN COALESCE(wp.tuesday_trend_detail, 'N/A')
                WHEN 2 THEN COALESCE(wp.wednesday_trend_detail, 'N/A')
                WHEN 3 THEN COALESCE(wp.thursday_trend_detail, 'N/A')
                WHEN 4 THEN COALESCE(wp.friday_trend_detail, 'N/A')
                WHEN 5 THEN COALESCE(wp.saturday_trend_detail, 'N/A')
                WHEN 6 THEN COALESCE(wp.sunday_trend_detail, 'N/A')
            END as '走势明细'
        FROM daily_data dd
        LEFT JOIN weekly_patterns wp ON (
            dd.symbol_id = wp.symbol_id 
            AND dd.trade_date >= DATE(wp.week_start) 
            AND dd.trade_date < DATE(wp.week_start, '+7 days')
        )
        WHERE dd.symbol_id = (SELECT id FROM symbols WHERE symbol = ?)
        AND wp.pattern IS NOT NULL
        AND CASE 
            WHEN dd.day_of_week = 0 THEN SUBSTR(wp.pattern, 1, 1)
            WHEN dd.day_of_week = 1 THEN SUBSTR(wp.pattern, 2, 1)
            WHEN dd.day_of_week = 2 THEN SUBSTR(wp.pattern, 3, 1)
            WHEN dd.day_of_week = 3 THEN SUBSTR(wp.pattern, 4, 1)
            WHEN dd.day_of_week = 4 THEN SUBSTR(wp.pattern, 5, 1)
            WHEN dd.day_of_week = 5 THEN SUBSTR(wp.pattern, 6, 1)
            WHEN dd.day_of_week = 6 THEN SUBSTR(wp.pattern, 7, 1)
        END IN ('A', 'M', 'D', 'X')
        ORDER BY dd.year, dd.trade_date
    """
    
    df = pd.read_sql_query(query, conn, params=(symbol_name,))
    return df


def create_statistics_report(conn):
    """创建统计报告"""
    print("=" * 60)
    print("周度模式与走势明细统计")
    print("=" * 60)
    
    excel_dir = os.path.join(REPORTS_DIR, 'excel')
    os.makedirs(excel_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = os.path.join(excel_dir, f'周度模式走势统计_{timestamp}.xlsx')
    
    # 走势明细的固定顺序
    trend_details = ['向上突破', '向下突破', '在区间内', '同时向上和向下突破']
    # 周度模式的固定顺序
    patterns = ['A', 'M', 'D', 'X']
    # 年份范围
    years = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # 处理每个交易对
        for symbol_name in ['BTCUSDT', 'ETHUSDT']:
            print(f"\n处理 {symbol_name}...")
            
            # 获取数据
            df = get_statistics_data(conn, symbol_name)
            
            if df.empty:
                print(f"  {symbol_name} 没有数据")
                continue
            
            # 创建统计表
            stats_data = []
            
            # 按年份统计
            for year in years:
                year_df = df[df['年份'] == year]
                if year_df.empty:
                    continue
                
                for pattern in patterns:
                    pattern_df = year_df[year_df['周度模式'] == pattern]
                    if pattern_df.empty:
                        continue
                    
                    row = {
                        '年份': year,
                        '周度模式': pattern,
                        '向上突破': len(pattern_df[pattern_df['走势明细'] == '向上突破']),
                        '向下突破': len(pattern_df[pattern_df['走势明细'] == '向下突破']),
                        '在区间内': len(pattern_df[pattern_df['走势明细'] == '在区间内']),
                        '同时向上和向下突破': len(pattern_df[pattern_df['走势明细'] == '同时向上和向下突破']),
                        '总计': len(pattern_df)
                    }
                    stats_data.append(row)
            
            # 添加总计行（所有年份）
            for pattern in patterns:
                pattern_df = df[df['周度模式'] == pattern]
                if pattern_df.empty:
                    continue
                
                row = {
                    '年份': '总计',
                    '周度模式': pattern,
                    '向上突破': len(pattern_df[pattern_df['走势明细'] == '向上突破']),
                    '向下突破': len(pattern_df[pattern_df['走势明细'] == '向下突破']),
                    '在区间内': len(pattern_df[pattern_df['走势明细'] == '在区间内']),
                    '同时向上和向下突破': len(pattern_df[pattern_df['走势明细'] == '同时向上和向下突破']),
                    '总计': len(pattern_df)
                }
                stats_data.append(row)
            
            # 创建DataFrame
            stats_df = pd.DataFrame(stats_data)
            
            # 写入Excel
            sheet_name = f'{symbol_name}_日统计'
            stats_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            style_excel_header(ws)
            style_data_cells(ws)
            auto_adjust_column_width(ws)
            
            print(f"  {symbol_name} 统计完成: {len(stats_df)} 行")
    
    print(f"\n统计报告已保存: {excel_path}")
    
    # 同时保存最新版本
    latest_path = os.path.join(excel_dir, '周度模式走势统计_最新.xlsx')
    import shutil
    shutil.copy(excel_path, latest_path)
    print(f"最新版本已保存: {latest_path}")
    
    print("\n" + "=" * 60)
    print("统计完成!")
    print("=" * 60)
    
    return excel_path


def main():
    """主函数"""
    conn = sqlite3.connect(DATABASE_PATH)
    try:
        create_statistics_report(conn)
    finally:
        conn.close()


if __name__ == '__main__':
    main()

