"""
周度模式数据导出脚本
导出周度模式（7字母模式）数据到Excel
"""

import sqlite3
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import DATABASE_PATH, REPORTS_DIR, TZ_UTC9

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def style_excel_header(ws, row_num=1):
    """设置Excel表头样式"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def style_data_cells(ws, start_row=2):
    """设置数据单元格样式"""
    data_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    xamdxam_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amdxamd_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border
            
            if cell.value == 'XAMDXAM':
                cell.fill = xamdxam_fill
            elif cell.value == 'AMDXAMD':
                cell.fill = amdxamd_fill


def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    for char in str(cell.value):
                        if '\u4e00' <= char <= '\u9fff':
                            cell_length += 1
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def export_weekly_patterns_to_excel(conn):
    """导出周度模式数据到Excel"""
    print("=" * 60)
    print("导出周度模式数据到Excel")
    print("=" * 60)
    
    excel_dir = os.path.join(REPORTS_DIR, 'excel')
    os.makedirs(excel_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = os.path.join(excel_dir, f'周度模式分析_{timestamp}.xlsx')
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # ========== 工作表1: 总体汇总 ==========
        print("生成工作表: 总体汇总...")
        query = """
            SELECT 
                s.symbol as '交易对',
                COUNT(*) as '总周数',
                SUM(CASE WHEN wp.pattern = 'XAMDXAM' THEN 1 ELSE 0 END) as 'XAMDXAM次数',
                SUM(CASE WHEN wp.pattern = 'AMDXAMD' THEN 1 ELSE 0 END) as 'AMDXAMD次数',
                ROUND(SUM(CASE WHEN wp.pattern = 'XAMDXAM' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMDXAM占比(%)',
                ROUND(SUM(CASE WHEN wp.pattern = 'AMDXAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDXAMD占比(%)',
                MIN(DATE(wp.week_start)) as '数据起始',
                MAX(DATE(wp.week_start)) as '数据结束'
            FROM weekly_patterns wp
            JOIN symbols s ON wp.symbol_id = s.id
            GROUP BY s.symbol
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='总体汇总', index=False)
        ws = writer.sheets['总体汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ========== 工作表2: 年度汇总 ==========
        print("生成工作表: 年度汇总...")
        query = """
            SELECT 
                s.symbol as '交易对',
                wp.year as '年份',
                COUNT(*) as '总周数',
                SUM(CASE WHEN wp.pattern = 'XAMDXAM' THEN 1 ELSE 0 END) as 'XAMDXAM次数',
                SUM(CASE WHEN wp.pattern = 'AMDXAMD' THEN 1 ELSE 0 END) as 'AMDXAMD次数',
                ROUND(SUM(CASE WHEN wp.pattern = 'XAMDXAM' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMDXAM占比(%)',
                ROUND(SUM(CASE WHEN wp.pattern = 'AMDXAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDXAMD占比(%)'
            FROM weekly_patterns wp
            JOIN symbols s ON wp.symbol_id = s.id
            GROUP BY s.symbol, wp.year
            ORDER BY s.symbol, wp.year
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='年度汇总', index=False)
        ws = writer.sheets['年度汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ========== 工作表3: BTC周度详细 ==========
        print("生成工作表: BTC周度详细...")
        query = """
            SELECT 
                DATE(wp.week_start) as '周开始日期',
                wp.year as '年份',
                wp.month as '月份',
                wp.week_of_year as '年内第几周',
                wp.pattern as '模式',
                wp.monday_trend_detail as '周一走势明细',
                ROUND(wp.monday_breakout_up_percent, 2) as '周一向上突破幅度(%)',
                ROUND(wp.monday_breakout_down_percent, 2) as '周一向下突破幅度(%)',
                wp.tuesday_trend_detail as '周二走势明细',
                ROUND(wp.tuesday_breakout_up_percent, 2) as '周二向上突破幅度(%)',
                ROUND(wp.tuesday_breakout_down_percent, 2) as '周二向下突破幅度(%)',
                wp.wednesday_trend_detail as '周三走势明细',
                ROUND(wp.wednesday_breakout_up_percent, 2) as '周三向上突破幅度(%)',
                ROUND(wp.wednesday_breakout_down_percent, 2) as '周三向下突破幅度(%)',
                wp.thursday_trend_detail as '周四走势明细',
                ROUND(wp.thursday_breakout_up_percent, 2) as '周四向上突破幅度(%)',
                ROUND(wp.thursday_breakout_down_percent, 2) as '周四向下突破幅度(%)',
                wp.friday_trend_detail as '周五走势明细',
                ROUND(wp.friday_breakout_up_percent, 2) as '周五向上突破幅度(%)',
                ROUND(wp.friday_breakout_down_percent, 2) as '周五向下突破幅度(%)',
                wp.saturday_trend_detail as '周六走势明细',
                ROUND(wp.saturday_breakout_up_percent, 2) as '周六向上突破幅度(%)',
                ROUND(wp.saturday_breakout_down_percent, 2) as '周六向下突破幅度(%)',
                wp.sunday_trend_detail as '周日走势明细',
                ROUND(wp.sunday_breakout_up_percent, 2) as '周日向上突破幅度(%)',
                ROUND(wp.sunday_breakout_down_percent, 2) as '周日向下突破幅度(%)'
            FROM weekly_patterns wp
            JOIN symbols s ON wp.symbol_id = s.id
            WHERE s.symbol = 'BTCUSDT'
            ORDER BY wp.week_start
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='BTC周度详细', index=False)
        ws = writer.sheets['BTC周度详细']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ========== 工作表4: ETH周度详细 ==========
        print("生成工作表: ETH周度详细...")
        query = """
            SELECT 
                DATE(wp.week_start) as '周开始日期',
                wp.year as '年份',
                wp.month as '月份',
                wp.week_of_year as '年内第几周',
                wp.pattern as '模式',
                wp.monday_trend_detail as '周一走势明细',
                ROUND(wp.monday_breakout_up_percent, 2) as '周一向上突破幅度(%)',
                ROUND(wp.monday_breakout_down_percent, 2) as '周一向下突破幅度(%)',
                wp.tuesday_trend_detail as '周二走势明细',
                ROUND(wp.tuesday_breakout_up_percent, 2) as '周二向上突破幅度(%)',
                ROUND(wp.tuesday_breakout_down_percent, 2) as '周二向下突破幅度(%)',
                wp.wednesday_trend_detail as '周三走势明细',
                ROUND(wp.wednesday_breakout_up_percent, 2) as '周三向上突破幅度(%)',
                ROUND(wp.wednesday_breakout_down_percent, 2) as '周三向下突破幅度(%)',
                wp.thursday_trend_detail as '周四走势明细',
                ROUND(wp.thursday_breakout_up_percent, 2) as '周四向上突破幅度(%)',
                ROUND(wp.thursday_breakout_down_percent, 2) as '周四向下突破幅度(%)',
                wp.friday_trend_detail as '周五走势明细',
                ROUND(wp.friday_breakout_up_percent, 2) as '周五向上突破幅度(%)',
                ROUND(wp.friday_breakout_down_percent, 2) as '周五向下突破幅度(%)',
                wp.saturday_trend_detail as '周六走势明细',
                ROUND(wp.saturday_breakout_up_percent, 2) as '周六向上突破幅度(%)',
                ROUND(wp.saturday_breakout_down_percent, 2) as '周六向下突破幅度(%)',
                wp.sunday_trend_detail as '周日走势明细',
                ROUND(wp.sunday_breakout_up_percent, 2) as '周日向上突破幅度(%)',
                ROUND(wp.sunday_breakout_down_percent, 2) as '周日向下突破幅度(%)'
            FROM weekly_patterns wp
            JOIN symbols s ON wp.symbol_id = s.id
            WHERE s.symbol = 'ETHUSDT'
            ORDER BY wp.week_start
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='ETH周度详细', index=False)
        ws = writer.sheets['ETH周度详细']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ========== 工作表5: 日数据(BTC) ==========
        print("生成工作表: BTC日数据...")
        query = """
            SELECT 
                trade_date as '日期',
                CASE day_of_week
                    WHEN 0 THEN '周一'
                    WHEN 1 THEN '周二'
                    WHEN 2 THEN '周三'
                    WHEN 3 THEN '周四'
                    WHEN 4 THEN '周五'
                    WHEN 5 THEN '周六'
                    WHEN 6 THEN '周日'
                END as '星期',
                day_high as '最高价',
                day_low as '最低价',
                day_open as '开盘价',
                day_close as '收盘价',
                day_volume as '成交量',
                data_quality_score as '数据质量分数'
            FROM daily_data
            WHERE symbol_id = (SELECT id FROM symbols WHERE symbol = 'BTCUSDT')
            ORDER BY trade_date
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='BTC日数据', index=False)
        ws = writer.sheets['BTC日数据']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ========== 工作表6: 日数据(ETH) ==========
        print("生成工作表: ETH日数据...")
        query = """
            SELECT 
                trade_date as '日期',
                CASE day_of_week
                    WHEN 0 THEN '周一'
                    WHEN 1 THEN '周二'
                    WHEN 2 THEN '周三'
                    WHEN 3 THEN '周四'
                    WHEN 4 THEN '周五'
                    WHEN 5 THEN '周六'
                    WHEN 6 THEN '周日'
                END as '星期',
                day_high as '最高价',
                day_low as '最低价',
                day_open as '开盘价',
                day_close as '收盘价',
                day_volume as '成交量',
                data_quality_score as '数据质量分数'
            FROM daily_data
            WHERE symbol_id = (SELECT id FROM symbols WHERE symbol = 'ETHUSDT')
            ORDER BY trade_date
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='ETH日数据', index=False)
        ws = writer.sheets['ETH日数据']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
    
    print(f"\n周度模式报告已保存: {excel_path}")
    
    # 同时保存最新版本
    latest_path = os.path.join(excel_dir, '周度模式分析_最新.xlsx')
    import shutil
    shutil.copy(excel_path, latest_path)
    print(f"最新版本已保存: {latest_path}")
    
    return excel_path


def main():
    """主函数"""
    conn = sqlite3.connect(DATABASE_PATH)
    
    try:
        # 检查数据是否存在
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM weekly_patterns")
        count = cursor.fetchone()[0]
        
        if count == 0:
            print("警告: 数据库中没有周度模式数据，请先运行:")
            print("  1. python scripts/fetch_daily_data.py")
            print("  2. python scripts/calculate_weekly_patterns.py")
            return
        
        print(f"数据库中有 {count} 条周度模式记录")
        
        export_weekly_patterns_to_excel(conn)
        
        print("\n" + "=" * 60)
        print("周度模式数据导出完成!")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        conn.close()


if __name__ == '__main__':
    main()

