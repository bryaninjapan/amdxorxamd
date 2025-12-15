"""
合并导出脚本
将月度模式分析和周度模式分析合并到同一个Excel文件
"""

import sqlite3
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import DATABASE_PATH, REPORTS_DIR, TZ_UTC9

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def style_excel_header(ws, row_num=1):
    """设置Excel表头样式"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def style_data_cells(ws, start_row=2):
    """设置数据单元格样式"""
    data_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 月度模式颜色
    amdx_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    xamd_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 周度模式颜色
    xamdxam_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amdxamd_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 月度模式样式
            if cell.value == 'AMDX':
                cell.fill = amdx_fill
            elif cell.value == 'XAMD':
                cell.fill = xamd_fill
            # 周度模式样式
            elif cell.value == 'XAMDXAM':
                cell.fill = xamdxam_fill
            elif cell.value == 'AMDXAMD':
                cell.fill = amdxamd_fill


def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    for char in str(cell.value):
                        if '\u4e00' <= char <= '\u9fff':
                            cell_length += 1
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def get_xamd_from_monthly_pattern(pattern_str, week_of_month):
    """
    根据月度模式表的pattern字符串和月内第几周
    转换为X/A/M/D走势
    """
    if 1 <= week_of_month <= 4 and len(pattern_str) == 4:
        return pattern_str[week_of_month - 1]
    return ''


def get_statistics_data(conn, symbol_name):
    """获取统计数据"""
    query = """
        SELECT 
            dd.year as '年份',
            CASE 
                WHEN wp.pattern IS NULL THEN 'N/A'
                WHEN dd.day_of_week = 0 THEN SUBSTR(wp.pattern, 1, 1)
                WHEN dd.day_of_week = 1 THEN SUBSTR(wp.pattern, 2, 1)
                WHEN dd.day_of_week = 2 THEN SUBSTR(wp.pattern, 3, 1)
                WHEN dd.day_of_week = 3 THEN SUBSTR(wp.pattern, 4, 1)
                WHEN dd.day_of_week = 4 THEN SUBSTR(wp.pattern, 5, 1)
                WHEN dd.day_of_week = 5 THEN SUBSTR(wp.pattern, 6, 1)
                WHEN dd.day_of_week = 6 THEN SUBSTR(wp.pattern, 7, 1)
                ELSE 'N/A'
            END as '周度模式',
            CASE dd.day_of_week
                WHEN 0 THEN COALESCE(wp.monday_trend_detail, 'N/A')
                WHEN 1 THEN COALESCE(wp.tuesday_trend_detail, 'N/A')
                WHEN 2 THEN COALESCE(wp.wednesday_trend_detail, 'N/A')
                WHEN 3 THEN COALESCE(wp.thursday_trend_detail, 'N/A')
                WHEN 4 THEN COALESCE(wp.friday_trend_detail, 'N/A')
                WHEN 5 THEN COALESCE(wp.saturday_trend_detail, 'N/A')
                WHEN 6 THEN COALESCE(wp.sunday_trend_detail, 'N/A')
            END as '走势明细'
        FROM daily_data dd
        LEFT JOIN weekly_patterns wp ON (
            dd.symbol_id = wp.symbol_id 
            AND dd.trade_date >= DATE(wp.week_start) 
            AND dd.trade_date < DATE(wp.week_start, '+7 days')
        )
        WHERE dd.symbol_id = (SELECT id FROM symbols WHERE symbol = ?)
        AND wp.pattern IS NOT NULL
        AND CASE 
            WHEN dd.day_of_week = 0 THEN SUBSTR(wp.pattern, 1, 1)
            WHEN dd.day_of_week = 1 THEN SUBSTR(wp.pattern, 2, 1)
            WHEN dd.day_of_week = 2 THEN SUBSTR(wp.pattern, 3, 1)
            WHEN dd.day_of_week = 3 THEN SUBSTR(wp.pattern, 4, 1)
            WHEN dd.day_of_week = 4 THEN SUBSTR(wp.pattern, 5, 1)
            WHEN dd.day_of_week = 5 THEN SUBSTR(wp.pattern, 6, 1)
            WHEN dd.day_of_week = 6 THEN SUBSTR(wp.pattern, 7, 1)
        END IN ('A', 'M', 'D', 'X')
        ORDER BY dd.year, dd.trade_date
    """
    
    df = pd.read_sql_query(query, conn, params=(symbol_name,))
    return df


def get_daily_data_with_pattern(conn, symbol_name):
    """获取日数据，包含日期、周度模式、走势明细、年份"""
    query = """
        SELECT 
            dd.trade_date as '日期',
            dd.year as '年份',
            CASE 
                WHEN wp.pattern IS NULL THEN 'N/A'
                WHEN dd.day_of_week = 0 THEN SUBSTR(wp.pattern, 1, 1)
                WHEN dd.day_of_week = 1 THEN SUBSTR(wp.pattern, 2, 1)
                WHEN dd.day_of_week = 2 THEN SUBSTR(wp.pattern, 3, 1)
                WHEN dd.day_of_week = 3 THEN SUBSTR(wp.pattern, 4, 1)
                WHEN dd.day_of_week = 4 THEN SUBSTR(wp.pattern, 5, 1)
                WHEN dd.day_of_week = 5 THEN SUBSTR(wp.pattern, 6, 1)
                WHEN dd.day_of_week = 6 THEN SUBSTR(wp.pattern, 7, 1)
                ELSE 'N/A'
            END as '周度模式',
            CASE dd.day_of_week
                WHEN 0 THEN COALESCE(wp.monday_trend_detail, 'N/A')
                WHEN 1 THEN COALESCE(wp.tuesday_trend_detail, 'N/A')
                WHEN 2 THEN COALESCE(wp.wednesday_trend_detail, 'N/A')
                WHEN 3 THEN COALESCE(wp.thursday_trend_detail, 'N/A')
                WHEN 4 THEN COALESCE(wp.friday_trend_detail, 'N/A')
                WHEN 5 THEN COALESCE(wp.saturday_trend_detail, 'N/A')
                WHEN 6 THEN COALESCE(wp.sunday_trend_detail, 'N/A')
            END as '走势明细'
        FROM daily_data dd
        LEFT JOIN weekly_patterns wp ON (
            dd.symbol_id = wp.symbol_id 
            AND dd.trade_date >= DATE(wp.week_start) 
            AND dd.trade_date < DATE(wp.week_start, '+7 days')
        )
        WHERE dd.symbol_id = (SELECT id FROM symbols WHERE symbol = ?)
        AND wp.pattern IS NOT NULL
        ORDER BY dd.trade_date
    """
    
    df = pd.read_sql_query(query, conn, params=(symbol_name,))
    df['日期'] = pd.to_datetime(df['日期'])
    return df


def calculate_consecutive_stats(df):
    """计算连续2天的统计"""
    from datetime import timedelta
    
    # 初始化统计字典（基础统计）
    stats = {
        'D_X_向上突破': {},      # D和X连续2天，都是"向上突破"
        'D_X_向下突破': {},      # D和X连续2天，都是"向下突破"
        'D_X_反方向': {},        # D→X连续反方向统计
        'D弱_X强': {},           # D弱→X强方向统计
        'X_A_在区间内': {},      # X→A连续，A为"在区间内"
        'X弱_A强': {}            # X弱→A强方向统计
    }
    
    # 初始化详细统计字典
    detailed_stats = {
        # D→X连续（同方向）
        'D上_X上': {},           # D为[向上突破]时，X为[向上突破]
        'D下_X下': {},           # D为[向下突破]时，X为[向下突破]
        
        # D→X连续反方向
        'D上_X下': {},           # D为[向上突破]时，X为[向下突破]
        'D上_X区间': {},         # D为[向上突破]时，X为[在区间内]
        'D下_X上': {},           # D为[向下突破]时，X为[向上突破]
        'D下_X区间': {},         # D为[向下突破]时，X为[在区间内]
        
        # D弱→X强
        'D区间_X上': {},         # D为[在区间内]时，X为[向上突破]
        'D区间_X下': {},         # D为[在区间内]时，X为[向下突破]
        
        # X→A连续
        'X上_A区间': {},         # X为"向上突破"时，A为"在区间内"
        'X下_A区间': {},         # X为"向下突破"时，A为"在区间内"
        'X双_A区间': {},         # X为"同时向上和向下突破"时，A为"在区间内"
        
        # X弱→A强
        'X上_A双': {},           # X为[向上突破]时，A为[同时向上和向下突破]
        'X下_A双': {}            # X为[向下突破]时，A为[同时向上和向下突破]
    }
    
    # 按日期排序
    df_sorted = df.sort_values('日期').reset_index(drop=True)
    
    # 遍历所有相邻的两天
    for i in range(len(df_sorted) - 1):
        today = df_sorted.iloc[i]
        tomorrow = df_sorted.iloc[i + 1]
        
        # 检查是否连续（相差1天）
        if (tomorrow['日期'] - today['日期']).days != 1:
            continue
        
        year = today['年份']
        today_pattern = today['周度模式']
        today_trend = today['走势明细']
        tomorrow_pattern = tomorrow['周度模式']
        tomorrow_trend = tomorrow['走势明细']
        
        # 统计1: D和X连续2天，都是"向上突破"
        if today_pattern == 'D' and tomorrow_pattern == 'X':
            if today_trend == '向上突破' and tomorrow_trend == '向上突破':
                if year not in stats['D_X_向上突破']:
                    stats['D_X_向上突破'][year] = 0
                stats['D_X_向上突破'][year] += 1
                # 详细统计
                if year not in detailed_stats['D上_X上']:
                    detailed_stats['D上_X上'][year] = 0
                detailed_stats['D上_X上'][year] += 1
        
        # 统计2: D和X连续2天，都是"向下突破"
        if today_pattern == 'D' and tomorrow_pattern == 'X':
            if today_trend == '向下突破' and tomorrow_trend == '向下突破':
                if year not in stats['D_X_向下突破']:
                    stats['D_X_向下突破'][year] = 0
                stats['D_X_向下突破'][year] += 1
                # 详细统计
                if year not in detailed_stats['D下_X下']:
                    detailed_stats['D下_X下'][year] = 0
                detailed_stats['D下_X下'][year] += 1
        
        # 统计3: D→X连续反方向统计（详细分类）
        if today_pattern == 'D' and tomorrow_pattern == 'X':
            if today_trend == '向上突破' and tomorrow_trend == '向下突破':
                if year not in stats['D_X_反方向']:
                    stats['D_X_反方向'][year] = 0
                stats['D_X_反方向'][year] += 1
                if year not in detailed_stats['D上_X下']:
                    detailed_stats['D上_X下'][year] = 0
                detailed_stats['D上_X下'][year] += 1
            elif today_trend == '向上突破' and tomorrow_trend == '在区间内':
                if year not in stats['D_X_反方向']:
                    stats['D_X_反方向'][year] = 0
                stats['D_X_反方向'][year] += 1
                if year not in detailed_stats['D上_X区间']:
                    detailed_stats['D上_X区间'][year] = 0
                detailed_stats['D上_X区间'][year] += 1
            elif today_trend == '向下突破' and tomorrow_trend == '向上突破':
                if year not in stats['D_X_反方向']:
                    stats['D_X_反方向'][year] = 0
                stats['D_X_反方向'][year] += 1
                if year not in detailed_stats['D下_X上']:
                    detailed_stats['D下_X上'][year] = 0
                detailed_stats['D下_X上'][year] += 1
            elif today_trend == '向下突破' and tomorrow_trend == '在区间内':
                if year not in stats['D_X_反方向']:
                    stats['D_X_反方向'][year] = 0
                stats['D_X_反方向'][year] += 1
                if year not in detailed_stats['D下_X区间']:
                    detailed_stats['D下_X区间'][year] = 0
                detailed_stats['D下_X区间'][year] += 1
        
        # 统计4: D弱→X强方向统计（详细分类）
        if today_pattern == 'D' and tomorrow_pattern == 'X':
            if today_trend == '在区间内' and tomorrow_trend == '向上突破':
                if year not in stats['D弱_X强']:
                    stats['D弱_X强'][year] = 0
                stats['D弱_X强'][year] += 1
                if year not in detailed_stats['D区间_X上']:
                    detailed_stats['D区间_X上'][year] = 0
                detailed_stats['D区间_X上'][year] += 1
            elif today_trend == '在区间内' and tomorrow_trend == '向下突破':
                if year not in stats['D弱_X强']:
                    stats['D弱_X强'][year] = 0
                stats['D弱_X强'][year] += 1
                if year not in detailed_stats['D区间_X下']:
                    detailed_stats['D区间_X下'][year] = 0
                detailed_stats['D区间_X下'][year] += 1
        
        # 统计5: X→A连续（详细分类）
        if today_pattern == 'X' and tomorrow_pattern == 'A':
            if today_trend == '向上突破' and tomorrow_trend == '在区间内':
                if year not in stats['X_A_在区间内']:
                    stats['X_A_在区间内'][year] = 0
                stats['X_A_在区间内'][year] += 1
                if year not in detailed_stats['X上_A区间']:
                    detailed_stats['X上_A区间'][year] = 0
                detailed_stats['X上_A区间'][year] += 1
            elif today_trend == '向下突破' and tomorrow_trend == '在区间内':
                if year not in stats['X_A_在区间内']:
                    stats['X_A_在区间内'][year] = 0
                stats['X_A_在区间内'][year] += 1
                if year not in detailed_stats['X下_A区间']:
                    detailed_stats['X下_A区间'][year] = 0
                detailed_stats['X下_A区间'][year] += 1
            elif today_trend == '同时向上和向下突破' and tomorrow_trend == '在区间内':
                if year not in stats['X_A_在区间内']:
                    stats['X_A_在区间内'][year] = 0
                stats['X_A_在区间内'][year] += 1
                if year not in detailed_stats['X双_A区间']:
                    detailed_stats['X双_A区间'][year] = 0
                detailed_stats['X双_A区间'][year] += 1
        
        # 统计6: X弱→A强方向统计（详细分类）
        if today_pattern == 'X' and tomorrow_pattern == 'A':
            if today_trend == '向上突破' and tomorrow_trend == '同时向上和向下突破':
                if year not in stats['X弱_A强']:
                    stats['X弱_A强'][year] = 0
                stats['X弱_A强'][year] += 1
                if year not in detailed_stats['X上_A双']:
                    detailed_stats['X上_A双'][year] = 0
                detailed_stats['X上_A双'][year] += 1
            elif today_trend == '向下突破' and tomorrow_trend == '同时向上和向下突破':
                if year not in stats['X弱_A强']:
                    stats['X弱_A强'][year] = 0
                stats['X弱_A强'][year] += 1
                if year not in detailed_stats['X下_A双']:
                    detailed_stats['X下_A双'][year] = 0
                detailed_stats['X下_A双'][year] += 1
    
    return stats, detailed_stats


def create_statistics_sheets(conn, writer):
    """创建统计工作表"""
    print("\n【日统计】")
    
    # 周度模式的固定顺序
    patterns = ['A', 'M', 'D', 'X']
    # 年份范围
    years = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
    
    # 处理每个交易对
    for symbol_name in ['BTCUSDT', 'ETHUSDT']:
        print(f"生成工作表: {symbol_name}_日统计...")
        
        # 获取基础统计数据
        df = get_statistics_data(conn, symbol_name)
        
        if df.empty:
            print(f"  {symbol_name} 没有数据")
            continue
        
        # 获取日数据用于连续统计
        df_daily = get_daily_data_with_pattern(conn, symbol_name)
        consecutive_stats, detailed_stats = calculate_consecutive_stats(df_daily)
        
        # 创建统计表
        stats_data = []
        
        # 按年份统计
        for year in years:
            year_df = df[df['年份'] == year]
            if year_df.empty:
                continue
            
            for pattern in patterns:
                pattern_df = year_df[year_df['周度模式'] == pattern]
                if pattern_df.empty:
                    continue
                
                row = {
                    '年份': year,
                    '周度模式': pattern,
                    '向上突破': len(pattern_df[pattern_df['走势明细'] == '向上突破']),
                    '向下突破': len(pattern_df[pattern_df['走势明细'] == '向下突破']),
                    '在区间内': len(pattern_df[pattern_df['走势明细'] == '在区间内']),
                    '同时向上和向下突破': len(pattern_df[pattern_df['走势明细'] == '同时向上和向下突破']),
                    '总计': len(pattern_df)
                }
                stats_data.append(row)
        
        # 添加总计行（所有年份）
        for pattern in patterns:
            pattern_df = df[df['周度模式'] == pattern]
            if pattern_df.empty:
                continue
            
            row = {
                '年份': '总计',
                '周度模式': pattern,
                '向上突破': len(pattern_df[pattern_df['走势明细'] == '向上突破']),
                '向下突破': len(pattern_df[pattern_df['走势明细'] == '向下突破']),
                '在区间内': len(pattern_df[pattern_df['走势明细'] == '在区间内']),
                '同时向上和向下突破': len(pattern_df[pattern_df['走势明细'] == '同时向上和向下突破']),
                '总计': len(pattern_df)
            }
            stats_data.append(row)
        
        # 添加连续统计行
        for year in years + ['总计']:
            if year == '总计':
                d_x_up_total = sum(consecutive_stats['D_X_向上突破'].values())
                d_x_down_total = sum(consecutive_stats['D_X_向下突破'].values())
                d_x_reverse_total = sum(consecutive_stats['D_X_反方向'].values())
                d_weak_x_strong_total = sum(consecutive_stats['D弱_X强'].values())
                x_a_total = sum(consecutive_stats['X_A_在区间内'].values())
                x_weak_a_strong_total = sum(consecutive_stats['X弱_A强'].values())
            else:
                d_x_up_total = consecutive_stats['D_X_向上突破'].get(year, 0)
                d_x_down_total = consecutive_stats['D_X_向下突破'].get(year, 0)
                d_x_reverse_total = consecutive_stats['D_X_反方向'].get(year, 0)
                d_weak_x_strong_total = consecutive_stats['D弱_X强'].get(year, 0)
                x_a_total = consecutive_stats['X_A_在区间内'].get(year, 0)
                x_weak_a_strong_total = consecutive_stats['X弱_A强'].get(year, 0)
            
            # D→X连续（同方向）
            if d_x_up_total > 0 or d_x_down_total > 0:
                row = {
                    '年份': year,
                    '周度模式': 'D→X连续',
                    '向上突破': d_x_up_total,
                    '向下突破': d_x_down_total,
                    '在区间内': 0,
                    '同时向上和向下突破': 0,
                    '总计': d_x_up_total + d_x_down_total
                }
                stats_data.append(row)
            
            # D→X连续反方向
            if d_x_reverse_total > 0:
                row = {
                    '年份': year,
                    '周度模式': 'D→X连续反方向',
                    '向上突破': 0,
                    '向下突破': 0,
                    '在区间内': d_x_reverse_total,
                    '同时向上和向下突破': 0,
                    '总计': d_x_reverse_total
                }
                stats_data.append(row)
            
            # D弱→X强
            if d_weak_x_strong_total > 0:
                row = {
                    '年份': year,
                    '周度模式': 'D弱→X强',
                    '向上突破': 0,
                    '向下突破': 0,
                    '在区间内': d_weak_x_strong_total,
                    '同时向上和向下突破': 0,
                    '总计': d_weak_x_strong_total
                }
                stats_data.append(row)
            
            # X→A连续（A在区间内）
            if x_a_total > 0:
                row = {
                    '年份': year,
                    '周度模式': 'X→A连续',
                    '向上突破': 0,
                    '向下突破': 0,
                    '在区间内': x_a_total,
                    '同时向上和向下突破': 0,
                    '总计': x_a_total
                }
                stats_data.append(row)
            
            # X弱→A强
            if x_weak_a_strong_total > 0:
                row = {
                    '年份': year,
                    '周度模式': 'X弱→A强',
                    '向上突破': 0,
                    '向下突破': 0,
                    '在区间内': 0,
                    '同时向上和向下突破': x_weak_a_strong_total,
                    '总计': x_weak_a_strong_total
                }
                stats_data.append(row)
        
        # 创建DataFrame
        stats_df = pd.DataFrame(stats_data)
        
        # 写入Excel
        sheet_name = f'{symbol_name}_日统计'
        stats_df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)


def create_detailed_consecutive_stats_sheets(conn, writer):
    """创建详细连续统计工作表"""
    print("\n【连续统计详细】")
    
    # 年份范围
    years = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
    
    # 处理每个交易对
    for symbol_name in ['BTCUSDT', 'ETHUSDT']:
        print(f"生成工作表: {symbol_name}_连续统计详细...")
        
        # 获取日数据用于连续统计
        df_daily = get_daily_data_with_pattern(conn, symbol_name)
        if df_daily.empty:
            print(f"  {symbol_name} 没有数据")
            continue
        
        _, detailed_stats = calculate_consecutive_stats(df_daily)
        
        # 创建详细统计表
        stats_data = []
        
        # 按年份统计
        for year in years:
            # D→X连续（同方向）
            d_up_x_up = detailed_stats['D上_X上'].get(year, 0)
            d_down_x_down = detailed_stats['D下_X下'].get(year, 0)
            if d_up_x_up > 0 or d_down_x_down > 0:
                stats_data.append({
                    '年份': year,
                    '统计类型': 'D→X连续',
                    '具体分类': 'D为[向上突破]时，X为[向上突破]',
                    '次数': d_up_x_up
                })
                stats_data.append({
                    '年份': year,
                    '统计类型': 'D→X连续',
                    '具体分类': 'D为[向下突破]时，X为[向下突破]',
                    '次数': d_down_x_down
                })
            
            # D→X连续反方向
            d_up_x_down = detailed_stats['D上_X下'].get(year, 0)
            d_up_x_range = detailed_stats['D上_X区间'].get(year, 0)
            d_down_x_up = detailed_stats['D下_X上'].get(year, 0)
            d_down_x_range = detailed_stats['D下_X区间'].get(year, 0)
            if d_up_x_down > 0 or d_up_x_range > 0 or d_down_x_up > 0 or d_down_x_range > 0:
                stats_data.append({
                    '年份': year,
                    '统计类型': 'D→X连续反方向',
                    '具体分类': 'D为[向上突破]时，X为[向下突破]',
                    '次数': d_up_x_down
                })
                stats_data.append({
                    '年份': year,
                    '统计类型': 'D→X连续反方向',
                    '具体分类': 'D为[向上突破]时，X为[在区间内]',
                    '次数': d_up_x_range
                })
                stats_data.append({
                    '年份': year,
                    '统计类型': 'D→X连续反方向',
                    '具体分类': 'D为[向下突破]时，X为[向上突破]',
                    '次数': d_down_x_up
                })
                stats_data.append({
                    '年份': year,
                    '统计类型': 'D→X连续反方向',
                    '具体分类': 'D为[向下突破]时，X为[在区间内]',
                    '次数': d_down_x_range
                })
            
            # D弱→X强
            d_range_x_up = detailed_stats['D区间_X上'].get(year, 0)
            d_range_x_down = detailed_stats['D区间_X下'].get(year, 0)
            d_weak_x_strong_total = d_range_x_up + d_range_x_down
            if d_weak_x_strong_total > 0:
                stats_data.append({
                    '年份': year,
                    '统计类型': 'D弱→X强',
                    '具体分类': 'D为[在区间内]时，X为[向上突破]或[向下突破]',
                    '次数': d_weak_x_strong_total
                })
            
            # X→A连续
            x_up_a_range = detailed_stats['X上_A区间'].get(year, 0)
            x_down_a_range = detailed_stats['X下_A区间'].get(year, 0)
            x_double_a_range = detailed_stats['X双_A区间'].get(year, 0)
            if x_up_a_range > 0 or x_down_a_range > 0 or x_double_a_range > 0:
                stats_data.append({
                    '年份': year,
                    '统计类型': 'X→A连续',
                    '具体分类': 'X为[向上突破]时，A为[在区间内]',
                    '次数': x_up_a_range
                })
                stats_data.append({
                    '年份': year,
                    '统计类型': 'X→A连续',
                    '具体分类': 'X为[向下突破]时，A为[在区间内]',
                    '次数': x_down_a_range
                })
                stats_data.append({
                    '年份': year,
                    '统计类型': 'X→A连续',
                    '具体分类': 'X为[同时向上和向下突破]时，A为[在区间内]',
                    '次数': x_double_a_range
                })
            
            # X弱→A强
            x_up_a_double = detailed_stats['X上_A双'].get(year, 0)
            x_down_a_double = detailed_stats['X下_A双'].get(year, 0)
            if x_up_a_double > 0 or x_down_a_double > 0:
                stats_data.append({
                    '年份': year,
                    '统计类型': 'X弱→A强',
                    '具体分类': 'X为[向上突破]时，A为[同时向上和向下突破]',
                    '次数': x_up_a_double
                })
                stats_data.append({
                    '年份': year,
                    '统计类型': 'X弱→A强',
                    '具体分类': 'X为[向下突破]时，A为[同时向上和向下突破]',
                    '次数': x_down_a_double
                })
        
        # 添加总计行
        # D→X连续（同方向）
        d_up_x_up_total = sum(detailed_stats['D上_X上'].values())
        d_down_x_down_total = sum(detailed_stats['D下_X下'].values())
        if d_up_x_up_total > 0 or d_down_x_down_total > 0:
            stats_data.append({
                '年份': '总计',
                '统计类型': 'D→X连续',
                '具体分类': 'D为[向上突破]时，X为[向上突破]',
                '次数': d_up_x_up_total
            })
            stats_data.append({
                '年份': '总计',
                '统计类型': 'D→X连续',
                '具体分类': 'D为[向下突破]时，X为[向下突破]',
                '次数': d_down_x_down_total
            })
        
        # D→X连续反方向
        d_up_x_down_total = sum(detailed_stats['D上_X下'].values())
        d_up_x_range_total = sum(detailed_stats['D上_X区间'].values())
        d_down_x_up_total = sum(detailed_stats['D下_X上'].values())
        d_down_x_range_total = sum(detailed_stats['D下_X区间'].values())
        if d_up_x_down_total > 0 or d_up_x_range_total > 0 or d_down_x_up_total > 0 or d_down_x_range_total > 0:
            stats_data.append({
                '年份': '总计',
                '统计类型': 'D→X连续反方向',
                '具体分类': 'D为[向上突破]时，X为[向下突破]',
                '次数': d_up_x_down_total
            })
            stats_data.append({
                '年份': '总计',
                '统计类型': 'D→X连续反方向',
                '具体分类': 'D为[向上突破]时，X为[在区间内]',
                '次数': d_up_x_range_total
            })
            stats_data.append({
                '年份': '总计',
                '统计类型': 'D→X连续反方向',
                '具体分类': 'D为[向下突破]时，X为[向上突破]',
                '次数': d_down_x_up_total
            })
            stats_data.append({
                '年份': '总计',
                '统计类型': 'D→X连续反方向',
                '具体分类': 'D为[向下突破]时，X为[在区间内]',
                '次数': d_down_x_range_total
            })
        
        # D弱→X强
        d_weak_x_strong_total = sum(detailed_stats['D区间_X上'].values()) + sum(detailed_stats['D区间_X下'].values())
        if d_weak_x_strong_total > 0:
            stats_data.append({
                '年份': '总计',
                '统计类型': 'D弱→X强',
                '具体分类': 'D为[在区间内]时，X为[向上突破]或[向下突破]',
                '次数': d_weak_x_strong_total
            })
        
        # X→A连续
        x_up_a_range_total = sum(detailed_stats['X上_A区间'].values())
        x_down_a_range_total = sum(detailed_stats['X下_A区间'].values())
        x_double_a_range_total = sum(detailed_stats['X双_A区间'].values())
        if x_up_a_range_total > 0 or x_down_a_range_total > 0 or x_double_a_range_total > 0:
            stats_data.append({
                '年份': '总计',
                '统计类型': 'X→A连续',
                '具体分类': 'X为[向上突破]时，A为[在区间内]',
                '次数': x_up_a_range_total
            })
            stats_data.append({
                '年份': '总计',
                '统计类型': 'X→A连续',
                '具体分类': 'X为[向下突破]时，A为[在区间内]',
                '次数': x_down_a_range_total
            })
            stats_data.append({
                '年份': '总计',
                '统计类型': 'X→A连续',
                '具体分类': 'X为[同时向上和向下突破]时，A为[在区间内]',
                '次数': x_double_a_range_total
            })
        
        # X弱→A强
        x_up_a_double_total = sum(detailed_stats['X上_A双'].values())
        x_down_a_double_total = sum(detailed_stats['X下_A双'].values())
        if x_up_a_double_total > 0 or x_down_a_double_total > 0:
            stats_data.append({
                '年份': '总计',
                '统计类型': 'X弱→A强',
                '具体分类': 'X为[向上突破]时，A为[同时向上和向下突破]',
                '次数': x_up_a_double_total
            })
            stats_data.append({
                '年份': '总计',
                '统计类型': 'X弱→A强',
                '具体分类': 'X为[向下突破]时，A为[同时向上和向下突破]',
                '次数': x_down_a_double_total
            })
        
        # 创建DataFrame
        stats_df = pd.DataFrame(stats_data)
        
        # 写入Excel
        sheet_name = f'{symbol_name}_连续统计详细'
        stats_df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)


def export_combined_report(conn):
    """导出合并报告（月度模式 + 周度模式）"""
    print("=" * 60)
    print("导出合并报告（月度模式 + 周度模式）")
    print("=" * 60)
    
    excel_dir = os.path.join(REPORTS_DIR, 'excel')
    os.makedirs(excel_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = os.path.join(excel_dir, f'完整分析报告_{timestamp}.xlsx')
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # ==================== 第一部分：月度模式分析 ====================
        print("\n【月度模式分析】")
        
        # 工作表1: 月度模式_总体汇总
        print("生成工作表: 月度模式_总体汇总...")
        query = """
            SELECT 
                s.symbol as '交易对',
                COUNT(*) as '总月数',
                SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
                SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
                ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
                ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)',
                MIN(mp.year || '-' || printf('%02d', mp.month)) as '数据起始月',
                MAX(mp.year || '-' || printf('%02d', mp.month)) as '数据结束月'
            FROM monthly_patterns mp
            JOIN symbols s ON mp.symbol_id = s.id
            GROUP BY s.symbol
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='月度模式_总体汇总', index=False)
        ws = writer.sheets['月度模式_总体汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表2: 月度模式_年度汇总
        print("生成工作表: 月度模式_年度汇总...")
        query = """
            SELECT 
                s.symbol as '交易对',
                mp.year as '年份',
                COUNT(*) as '总月数',
                SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
                SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
                ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
                ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)',
                SUM(CASE WHEN mp.is_breakout_up = 1 THEN 1 ELSE 0 END) as '向上突破次数',
                SUM(CASE WHEN mp.is_breakout_down = 1 THEN 1 ELSE 0 END) as '向下突破次数',
                ROUND(AVG(CASE WHEN mp.breakout_up_percent IS NOT NULL THEN mp.breakout_up_percent END), 2) as '平均向上突破幅度(%)',
                ROUND(AVG(CASE WHEN mp.breakout_down_percent IS NOT NULL THEN mp.breakout_down_percent END), 2) as '平均向下突破幅度(%)'
            FROM monthly_patterns mp
            JOIN symbols s ON mp.symbol_id = s.id
            GROUP BY s.symbol, mp.year
            ORDER BY s.symbol, mp.year
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='月度模式_年度汇总', index=False)
        ws = writer.sheets['月度模式_年度汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表3: BTC月份分布统计
        print("生成工作表: BTC月份分布统计...")
        query = """
            SELECT 
                mp.month as '月份',
                COUNT(*) as '总月数',
                SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
                SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
                ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
                ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)'
            FROM monthly_patterns mp
            JOIN symbols s ON mp.symbol_id = s.id
            WHERE s.symbol = 'BTCUSDT'
            GROUP BY mp.month
            ORDER BY mp.month
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='BTC月份分布统计', index=False)
        ws = writer.sheets['BTC月份分布统计']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表4: ETH月份分布统计
        print("生成工作表: ETH月份分布统计...")
        query = """
            SELECT 
                mp.month as '月份',
                COUNT(*) as '总月数',
                SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1 ELSE 0 END) as 'AMDX次数',
                SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1 ELSE 0 END) as 'XAMD次数',
                ROUND(SUM(CASE WHEN mp.pattern = 'AMDX' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDX占比(%)',
                ROUND(SUM(CASE WHEN mp.pattern = 'XAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMD占比(%)'
            FROM monthly_patterns mp
            JOIN symbols s ON mp.symbol_id = s.id
            WHERE s.symbol = 'ETHUSDT'
            GROUP BY mp.month
            ORDER BY mp.month
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='ETH月份分布统计', index=False)
        ws = writer.sheets['ETH月份分布统计']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表5-6: BTCUSDT_周数据 和 ETHUSDT_周数据（带走势列）
        symbols_query = "SELECT id, symbol FROM symbols WHERE is_active = 1"
        symbols_df = pd.read_sql_query(symbols_query, conn)
        
        monthly_patterns_query = """
            SELECT 
                symbol_id,
                year,
                month,
                pattern
            FROM monthly_patterns
        """
        monthly_patterns_df = pd.read_sql_query(monthly_patterns_query, conn)
        
        for _, row in symbols_df.iterrows():
            symbol_id = row['id']
            symbol = row['symbol']
            
            print(f"生成工作表: {symbol}_周数据...")
            query = """
                SELECT 
                    week_start as '周开始时间',
                    week_end as '周结束时间',
                    year as '年份',
                    month as '月份',
                    week_of_year as '年内第几周',
                    week_of_month as '月内第几周',
                    week_high as '周最高价',
                    week_low as '周最低价',
                    week_open as '周开盘价',
                    week_close as '周收盘价',
                    data_points as '数据点数',
                    data_quality_score as '数据质量分数'
                FROM weekly_data
                WHERE symbol_id = ?
                ORDER BY week_start
            """
            df = pd.read_sql_query(query, conn, params=(symbol_id,))
            
            # 根据月度模式表计算X/A/M/D走势
            xamd_patterns = []
            for idx, row_data in df.iterrows():
                week_of_month = row_data['月内第几周']
                year = row_data['年份']
                month = row_data['月份']
                if 1 <= week_of_month <= 4:
                    monthly_pattern = monthly_patterns_df[
                        (monthly_patterns_df['symbol_id'] == symbol_id) &
                        (monthly_patterns_df['year'] == year) &
                        (monthly_patterns_df['month'] == month)
                    ]
                    if len(monthly_pattern) > 0:
                        pattern_str = monthly_pattern.iloc[0]['pattern']
                        xamd_patterns.append(get_xamd_from_monthly_pattern(pattern_str, week_of_month))
                    else:
                        xamd_patterns.append('')
                else:
                    xamd_patterns.append('')
            
            df.insert(df.columns.get_loc('月内第几周') + 1, '走势', xamd_patterns)
            df.to_excel(writer, sheet_name=f'{symbol}_周数据', index=False)
            ws = writer.sheets[f'{symbol}_周数据']
            style_excel_header(ws)
            style_data_cells(ws)
            auto_adjust_column_width(ws)
        
        # ==================== 第二部分：周度模式分析 ====================
        print("\n【周度模式分析】")
        
        # 工作表: 周度模式_总体汇总
        print("生成工作表: 周度模式_总体汇总...")
        query = """
            SELECT 
                s.symbol as '交易对',
                COUNT(*) as '总周数',
                SUM(CASE WHEN wp.pattern = 'XAMDXAM' THEN 1 ELSE 0 END) as 'XAMDXAM次数',
                SUM(CASE WHEN wp.pattern = 'AMDXAMD' THEN 1 ELSE 0 END) as 'AMDXAMD次数',
                ROUND(SUM(CASE WHEN wp.pattern = 'XAMDXAM' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMDXAM占比(%)',
                ROUND(SUM(CASE WHEN wp.pattern = 'AMDXAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDXAMD占比(%)',
                MIN(DATE(wp.week_start)) as '数据起始',
                MAX(DATE(wp.week_start)) as '数据结束'
            FROM weekly_patterns wp
            JOIN symbols s ON wp.symbol_id = s.id
            GROUP BY s.symbol
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='周度模式_总体汇总', index=False)
        ws = writer.sheets['周度模式_总体汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表: 周度模式_年度汇总
        print("生成工作表: 周度模式_年度汇总...")
        query = """
            SELECT 
                s.symbol as '交易对',
                wp.year as '年份',
                COUNT(*) as '总周数',
                SUM(CASE WHEN wp.pattern = 'XAMDXAM' THEN 1 ELSE 0 END) as 'XAMDXAM次数',
                SUM(CASE WHEN wp.pattern = 'AMDXAMD' THEN 1 ELSE 0 END) as 'AMDXAMD次数',
                ROUND(SUM(CASE WHEN wp.pattern = 'XAMDXAM' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'XAMDXAM占比(%)',
                ROUND(SUM(CASE WHEN wp.pattern = 'AMDXAMD' THEN 1.0 ELSE 0 END) * 100.0 / COUNT(*), 1) as 'AMDXAMD占比(%)'
            FROM weekly_patterns wp
            JOIN symbols s ON wp.symbol_id = s.id
            GROUP BY s.symbol, wp.year
            ORDER BY s.symbol, wp.year
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='周度模式_年度汇总', index=False)
        ws = writer.sheets['周度模式_年度汇总']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表: BTC周度详细
        print("生成工作表: BTC周度详细...")
        query = """
            SELECT 
                DATE(wp.week_start) as '周开始日期',
                wp.year as '年份',
                wp.month as '月份',
                wp.week_of_year as '年内第几周',
                wp.pattern as '模式',
                wp.monday_trend_detail as '周一走势明细',
                ROUND(wp.monday_breakout_up_percent, 2) as '周一向上突破幅度(%)',
                ROUND(wp.monday_breakout_down_percent, 2) as '周一向下突破幅度(%)',
                wp.tuesday_trend_detail as '周二走势明细',
                ROUND(wp.tuesday_breakout_up_percent, 2) as '周二向上突破幅度(%)',
                ROUND(wp.tuesday_breakout_down_percent, 2) as '周二向下突破幅度(%)',
                wp.wednesday_trend_detail as '周三走势明细',
                ROUND(wp.wednesday_breakout_up_percent, 2) as '周三向上突破幅度(%)',
                ROUND(wp.wednesday_breakout_down_percent, 2) as '周三向下突破幅度(%)',
                wp.thursday_trend_detail as '周四走势明细',
                ROUND(wp.thursday_breakout_up_percent, 2) as '周四向上突破幅度(%)',
                ROUND(wp.thursday_breakout_down_percent, 2) as '周四向下突破幅度(%)',
                wp.friday_trend_detail as '周五走势明细',
                ROUND(wp.friday_breakout_up_percent, 2) as '周五向上突破幅度(%)',
                ROUND(wp.friday_breakout_down_percent, 2) as '周五向下突破幅度(%)',
                wp.saturday_trend_detail as '周六走势明细',
                ROUND(wp.saturday_breakout_up_percent, 2) as '周六向上突破幅度(%)',
                ROUND(wp.saturday_breakout_down_percent, 2) as '周六向下突破幅度(%)',
                wp.sunday_trend_detail as '周日走势明细',
                ROUND(wp.sunday_breakout_up_percent, 2) as '周日向上突破幅度(%)',
                ROUND(wp.sunday_breakout_down_percent, 2) as '周日向下突破幅度(%)'
            FROM weekly_patterns wp
            JOIN symbols s ON wp.symbol_id = s.id
            WHERE s.symbol = 'BTCUSDT'
            ORDER BY wp.week_start
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='BTC周度详细', index=False)
        ws = writer.sheets['BTC周度详细']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表: ETH周度详细
        print("生成工作表: ETH周度详细...")
        query = """
            SELECT 
                DATE(wp.week_start) as '周开始日期',
                wp.year as '年份',
                wp.month as '月份',
                wp.week_of_year as '年内第几周',
                wp.pattern as '模式',
                wp.monday_trend_detail as '周一走势明细',
                ROUND(wp.monday_breakout_up_percent, 2) as '周一向上突破幅度(%)',
                ROUND(wp.monday_breakout_down_percent, 2) as '周一向下突破幅度(%)',
                wp.tuesday_trend_detail as '周二走势明细',
                ROUND(wp.tuesday_breakout_up_percent, 2) as '周二向上突破幅度(%)',
                ROUND(wp.tuesday_breakout_down_percent, 2) as '周二向下突破幅度(%)',
                wp.wednesday_trend_detail as '周三走势明细',
                ROUND(wp.wednesday_breakout_up_percent, 2) as '周三向上突破幅度(%)',
                ROUND(wp.wednesday_breakout_down_percent, 2) as '周三向下突破幅度(%)',
                wp.thursday_trend_detail as '周四走势明细',
                ROUND(wp.thursday_breakout_up_percent, 2) as '周四向上突破幅度(%)',
                ROUND(wp.thursday_breakout_down_percent, 2) as '周四向下突破幅度(%)',
                wp.friday_trend_detail as '周五走势明细',
                ROUND(wp.friday_breakout_up_percent, 2) as '周五向上突破幅度(%)',
                ROUND(wp.friday_breakout_down_percent, 2) as '周五向下突破幅度(%)',
                wp.saturday_trend_detail as '周六走势明细',
                ROUND(wp.saturday_breakout_up_percent, 2) as '周六向上突破幅度(%)',
                ROUND(wp.saturday_breakout_down_percent, 2) as '周六向下突破幅度(%)',
                wp.sunday_trend_detail as '周日走势明细',
                ROUND(wp.sunday_breakout_up_percent, 2) as '周日向上突破幅度(%)',
                ROUND(wp.sunday_breakout_down_percent, 2) as '周日向下突破幅度(%)'
            FROM weekly_patterns wp
            JOIN symbols s ON wp.symbol_id = s.id
            WHERE s.symbol = 'ETHUSDT'
            ORDER BY wp.week_start
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='ETH周度详细', index=False)
        ws = writer.sheets['ETH周度详细']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表: BTC日数据
        print("生成工作表: BTC日数据...")
        query = """
            SELECT 
                dd.trade_date as '日期',
                CASE dd.day_of_week
                    WHEN 0 THEN '周一'
                    WHEN 1 THEN '周二'
                    WHEN 2 THEN '周三'
                    WHEN 3 THEN '周四'
                    WHEN 4 THEN '周五'
                    WHEN 5 THEN '周六'
                    WHEN 6 THEN '周日'
                END as '星期',
                CASE 
                    WHEN wp.pattern IS NULL THEN 'N/A'
                    WHEN dd.day_of_week = 0 THEN SUBSTR(wp.pattern, 1, 1)  -- 周一，第1个字母
                    WHEN dd.day_of_week = 1 THEN SUBSTR(wp.pattern, 2, 1)  -- 周二，第2个字母
                    WHEN dd.day_of_week = 2 THEN SUBSTR(wp.pattern, 3, 1)  -- 周三，第3个字母
                    WHEN dd.day_of_week = 3 THEN SUBSTR(wp.pattern, 4, 1)  -- 周四，第4个字母
                    WHEN dd.day_of_week = 4 THEN SUBSTR(wp.pattern, 5, 1)  -- 周五，第5个字母
                    WHEN dd.day_of_week = 5 THEN SUBSTR(wp.pattern, 6, 1)  -- 周六，第6个字母
                    WHEN dd.day_of_week = 6 THEN SUBSTR(wp.pattern, 7, 1)  -- 周日，第7个字母
                    ELSE 'N/A'
                END as '周度模式',
                CASE dd.day_of_week
                    WHEN 0 THEN COALESCE(wp.monday_trend_detail, 'N/A')
                    WHEN 1 THEN COALESCE(wp.tuesday_trend_detail, 'N/A')
                    WHEN 2 THEN COALESCE(wp.wednesday_trend_detail, 'N/A')
                    WHEN 3 THEN COALESCE(wp.thursday_trend_detail, 'N/A')
                    WHEN 4 THEN COALESCE(wp.friday_trend_detail, 'N/A')
                    WHEN 5 THEN COALESCE(wp.saturday_trend_detail, 'N/A')
                    WHEN 6 THEN COALESCE(wp.sunday_trend_detail, 'N/A')
                END as '走势明细',
                CASE dd.day_of_week
                    WHEN 0 THEN ROUND(wp.monday_breakout_up_percent, 2)
                    WHEN 1 THEN ROUND(wp.tuesday_breakout_up_percent, 2)
                    WHEN 2 THEN ROUND(wp.wednesday_breakout_up_percent, 2)
                    WHEN 3 THEN ROUND(wp.thursday_breakout_up_percent, 2)
                    WHEN 4 THEN ROUND(wp.friday_breakout_up_percent, 2)
                    WHEN 5 THEN ROUND(wp.saturday_breakout_up_percent, 2)
                    WHEN 6 THEN ROUND(wp.sunday_breakout_up_percent, 2)
                END as '向上突破幅度(%)',
                CASE dd.day_of_week
                    WHEN 0 THEN ROUND(wp.monday_breakout_down_percent, 2)
                    WHEN 1 THEN ROUND(wp.tuesday_breakout_down_percent, 2)
                    WHEN 2 THEN ROUND(wp.wednesday_breakout_down_percent, 2)
                    WHEN 3 THEN ROUND(wp.thursday_breakout_down_percent, 2)
                    WHEN 4 THEN ROUND(wp.friday_breakout_down_percent, 2)
                    WHEN 5 THEN ROUND(wp.saturday_breakout_down_percent, 2)
                    WHEN 6 THEN ROUND(wp.sunday_breakout_down_percent, 2)
                END as '向下突破幅度(%)',
                dd.day_high as '最高价',
                dd.day_low as '最低价',
                dd.day_open as '开盘价',
                dd.day_close as '收盘价',
                dd.day_volume as '成交量',
                dd.data_quality_score as '数据质量分数'
            FROM daily_data dd
            LEFT JOIN weekly_patterns wp ON (
                dd.symbol_id = wp.symbol_id 
                AND dd.trade_date >= DATE(wp.week_start) 
                AND dd.trade_date < DATE(wp.week_start, '+7 days')
            )
            WHERE dd.symbol_id = (SELECT id FROM symbols WHERE symbol = 'BTCUSDT')
            ORDER BY dd.trade_date
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='BTC日数据', index=False)
        ws = writer.sheets['BTC日数据']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # 工作表: ETH日数据
        print("生成工作表: ETH日数据...")
        query = """
            SELECT 
                dd.trade_date as '日期',
                CASE dd.day_of_week
                    WHEN 0 THEN '周一'
                    WHEN 1 THEN '周二'
                    WHEN 2 THEN '周三'
                    WHEN 3 THEN '周四'
                    WHEN 4 THEN '周五'
                    WHEN 5 THEN '周六'
                    WHEN 6 THEN '周日'
                END as '星期',
                CASE 
                    WHEN wp.pattern IS NULL THEN 'N/A'
                    WHEN dd.day_of_week = 0 THEN SUBSTR(wp.pattern, 1, 1)  -- 周一，第1个字母
                    WHEN dd.day_of_week = 1 THEN SUBSTR(wp.pattern, 2, 1)  -- 周二，第2个字母
                    WHEN dd.day_of_week = 2 THEN SUBSTR(wp.pattern, 3, 1)  -- 周三，第3个字母
                    WHEN dd.day_of_week = 3 THEN SUBSTR(wp.pattern, 4, 1)  -- 周四，第4个字母
                    WHEN dd.day_of_week = 4 THEN SUBSTR(wp.pattern, 5, 1)  -- 周五，第5个字母
                    WHEN dd.day_of_week = 5 THEN SUBSTR(wp.pattern, 6, 1)  -- 周六，第6个字母
                    WHEN dd.day_of_week = 6 THEN SUBSTR(wp.pattern, 7, 1)  -- 周日，第7个字母
                    ELSE 'N/A'
                END as '周度模式',
                CASE dd.day_of_week
                    WHEN 0 THEN COALESCE(wp.monday_trend_detail, 'N/A')
                    WHEN 1 THEN COALESCE(wp.tuesday_trend_detail, 'N/A')
                    WHEN 2 THEN COALESCE(wp.wednesday_trend_detail, 'N/A')
                    WHEN 3 THEN COALESCE(wp.thursday_trend_detail, 'N/A')
                    WHEN 4 THEN COALESCE(wp.friday_trend_detail, 'N/A')
                    WHEN 5 THEN COALESCE(wp.saturday_trend_detail, 'N/A')
                    WHEN 6 THEN COALESCE(wp.sunday_trend_detail, 'N/A')
                END as '走势明细',
                CASE dd.day_of_week
                    WHEN 0 THEN ROUND(wp.monday_breakout_up_percent, 2)
                    WHEN 1 THEN ROUND(wp.tuesday_breakout_up_percent, 2)
                    WHEN 2 THEN ROUND(wp.wednesday_breakout_up_percent, 2)
                    WHEN 3 THEN ROUND(wp.thursday_breakout_up_percent, 2)
                    WHEN 4 THEN ROUND(wp.friday_breakout_up_percent, 2)
                    WHEN 5 THEN ROUND(wp.saturday_breakout_up_percent, 2)
                    WHEN 6 THEN ROUND(wp.sunday_breakout_up_percent, 2)
                END as '向上突破幅度(%)',
                CASE dd.day_of_week
                    WHEN 0 THEN ROUND(wp.monday_breakout_down_percent, 2)
                    WHEN 1 THEN ROUND(wp.tuesday_breakout_down_percent, 2)
                    WHEN 2 THEN ROUND(wp.wednesday_breakout_down_percent, 2)
                    WHEN 3 THEN ROUND(wp.thursday_breakout_down_percent, 2)
                    WHEN 4 THEN ROUND(wp.friday_breakout_down_percent, 2)
                    WHEN 5 THEN ROUND(wp.saturday_breakout_down_percent, 2)
                    WHEN 6 THEN ROUND(wp.sunday_breakout_down_percent, 2)
                END as '向下突破幅度(%)',
                dd.day_high as '最高价',
                dd.day_low as '最低价',
                dd.day_open as '开盘价',
                dd.day_close as '收盘价',
                dd.day_volume as '成交量',
                dd.data_quality_score as '数据质量分数'
            FROM daily_data dd
            LEFT JOIN weekly_patterns wp ON (
                dd.symbol_id = wp.symbol_id 
                AND dd.trade_date >= DATE(wp.week_start) 
                AND dd.trade_date < DATE(wp.week_start, '+7 days')
            )
            WHERE dd.symbol_id = (SELECT id FROM symbols WHERE symbol = 'ETHUSDT')
            ORDER BY dd.trade_date
        """
        df = pd.read_sql_query(query, conn)
        df.to_excel(writer, sheet_name='ETH日数据', index=False)
        ws = writer.sheets['ETH日数据']
        style_excel_header(ws)
        style_data_cells(ws)
        auto_adjust_column_width(ws)
        
        # ==================== 第三部分：日统计 ====================
        create_statistics_sheets(conn, writer)
        
        # ==================== 第四部分：连续统计详细 ====================
        create_detailed_consecutive_stats_sheets(conn, writer)
    
    print(f"\n合并报告已保存: {excel_path}")
    
    # 同时保存最新版本
    latest_path = os.path.join(excel_dir, '完整分析报告_最新.xlsx')
    import shutil
    shutil.copy(excel_path, latest_path)
    print(f"最新版本已保存: {latest_path}")
    
    print("\n" + "=" * 60)
    print("合并报告导出完成!")
    print("=" * 60)
    
    return excel_path


def main():
    """主函数"""
    conn = sqlite3.connect(DATABASE_PATH)
    try:
        export_combined_report(conn)
    finally:
        conn.close()


if __name__ == '__main__':
    main()

